"""
Timesheet Excel Generator
=======================
Stateless — no database access, no filesystem writes.
Returns raw Excel bytes directly to the caller.

Public API
----------
- build_excel_filename(candidate_name, start, end) -> str
- generate_timesheet_excel(timesheet, candidate_name) -> bytes
"""

from __future__ import annotations

import re
import logging
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.core.db.models import Timesheet
from app.services.timesheet_pdf import derive_timesheet_type, _sanitise

logger = logging.getLogger(__name__)

# Brand styles
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin', color='E4E4E7'),
    right=Side(style='thin', color='E4E4E7'),
    top=Side(style='thin', color='E4E4E7'),
    bottom=Side(style='thin', color='E4E4E7')
)

def build_excel_filename(candidate_name: str, start: datetime, end: datetime) -> str:
    """
    Generate a human-readable Excel filename.
    Example: timesheet_Jane_Doe_2025-07-21_2025-07-27.xlsx
    """
    name      = _sanitise(candidate_name)
    start_str = start.strftime("%Y-%m-%d") if hasattr(start, "strftime") else str(start)
    end_str   = end.strftime("%Y-%m-%d")   if hasattr(end,   "strftime") else str(end)
    return f"timesheet_{name}_{start_str}_{end_str}.xlsx"

def generate_timesheet_excel(timesheet: Timesheet, candidate_name: str) -> bytes:
    """
    Build a structured Excel file for the given timesheet.
    Returns raw bytes. Never writes to disk or database.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    period_type  = derive_timesheet_type(
        timesheet.period_start_date, timesheet.period_end_date
    )
    period_label = (
        f"{timesheet.period_start_date.strftime('%d %b %Y')}"
        f" - "
        f"{timesheet.period_end_date.strftime('%d %b %Y')}"
    )

    # Title
    ws["A1"] = "Lorvish Platform - Timesheet Submission"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="4F46E5")
    ws.merge_cells("A1:C1")

    # Summary
    summary_data = [
        ("Candidate", candidate_name),
        ("Period", period_label),
        ("Type", period_type),
        ("Total Hours", f"{timesheet.total_hours}h"),
        ("Status", "Submitted")
    ]
    
    current_row = 3
    for k, v in summary_data:
        ws.cell(row=current_row, column=1, value=k).font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=v).font = NORMAL_FONT
        current_row += 1

    current_row += 1

    # Daily Entries Header
    ws.cell(row=current_row, column=1, value="Date").fill = HEADER_FILL
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
    
    ws.cell(row=current_row, column=2, value="Hours").fill = HEADER_FILL
    ws.cell(row=current_row, column=2).font = HEADER_FONT
    ws.cell(row=current_row, column=2).alignment = CENTER_ALIGN
    
    ws.cell(row=current_row, column=3, value="Task Description").fill = HEADER_FILL
    ws.cell(row=current_row, column=3).font = HEADER_FONT
    ws.cell(row=current_row, column=3).alignment = LEFT_ALIGN
    
    current_row += 1

    # Daily Entries Data
    sorted_entries = sorted(timesheet.entries, key=lambda e: e.date)
    if sorted_entries:
        for entry in sorted_entries:
            date_str = (
                entry.date.strftime("%a, %d %b %Y")
                if hasattr(entry.date, "strftime") else str(entry.date)
            )
            
            c1 = ws.cell(row=current_row, column=1, value=date_str)
            c2 = ws.cell(row=current_row, column=2, value=entry.hours_worked)
            c3 = ws.cell(row=current_row, column=3, value=entry.task_description)
            
            c1.font = NORMAL_FONT
            c2.font = NORMAL_FONT
            c3.font = NORMAL_FONT
            
            c1.border = THIN_BORDER
            c2.border = THIN_BORDER
            c3.border = THIN_BORDER
            
            c1.alignment = LEFT_ALIGN
            c2.alignment = CENTER_ALIGN
            c3.alignment = LEFT_ALIGN
            
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="No daily entries recorded.")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1

    current_row += 1
    
    # Total Hours Footer
    ws.cell(row=current_row, column=1, value="Total Hours").font = BOLD_FONT
    ws.cell(row=current_row, column=2, value=timesheet.total_hours).font = BOLD_FONT
    ws.cell(row=current_row, column=2).alignment = CENTER_ALIGN
    current_row += 2

    # Notes
    if timesheet.notes:
        ws.cell(row=current_row, column=1, value="Notes").font = BOLD_FONT
        current_row += 1
        ws.cell(row=current_row, column=1, value=timesheet.notes).font = NORMAL_FONT
        current_row += 2

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
